import os
import datetime
import openpyxl
from openpyxl.utils import get_column_letter

# Paths
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_EXCEL_PATH = r"c:\projetos\finscript\Dashboard Ocimar 2023.xlsx"

class ExcelManager:
    def __init__(self, filepath=DEFAULT_EXCEL_PATH):
        self.filepath = filepath
        if not os.path.exists(self.filepath):
            raise FileNotFoundError(f"Planilha base não encontrada em: {self.filepath}")

    def get_workbook(self, read_only=False, data_only=False):
        return openpyxl.load_workbook(self.filepath, read_only=read_only, data_only=data_only)

    def _clean_dcto(self, val):
        """Evita que números de documento formatados como Data no Excel sejam convertidos para datetime"""
        if isinstance(val, (datetime.datetime, datetime.date)):
            # Excel base date é 1899-12-30 devido ao bug do ano bissexto de 1900
            base_date = datetime.date(1899, 12, 30)
            if isinstance(val, datetime.datetime):
                val = val.date()
            return (val - base_date).days
        return val


    def get_categories(self):
        """
        Retorna a estrutura do Plano de Contas:
        { "GRUPO": ["SUBGRUPO1", "SUBGRUPO2", ...] }
        """
        wb = self.get_workbook(data_only=True)
        sheet = wb["Plano de contas"]
        categories = {}
        
        # O cabeçalho está na linha 6, e os dados abaixo dela. 
        # Colunas com Grupos ocorrem a cada 2 colunas (colunas pares B, D, F, H, J, L, N)
        for col in range(2, sheet.max_column + 1, 2):
            group_name = sheet.cell(row=6, column=col).value
            if group_name:
                subgroups = []
                for r in range(7, sheet.max_row + 1):
                    sub_val = sheet.cell(row=r, column=col).value
                    if sub_val is not None and str(sub_val).strip() != '':
                        subgroups.append(str(sub_val).strip())
                # Armazena mapeamento
                categories[group_name.strip()] = subgroups
                
        wb.close()
        return categories

    def _parse_date(self, val):
        """Converte valores de data do Excel para string YYYY-MM-DD"""
        if isinstance(val, (datetime.datetime, datetime.date)):
            return val.strftime("%Y-%m-%d")
        if isinstance(val, str):
            # Limpa possíveis formatações
            val = val.strip().split(" ")[0]
            try:
                # valida se está no formato YYYY-MM-DD
                datetime.datetime.strptime(val, "%Y-%m-%d")
                return val
            except ValueError:
                pass
        return None

    def _to_date_obj(self, date_str):
        """Converte string YYYY-MM-DD para objeto datetime.date"""
        if not date_str:
            return None
        if isinstance(date_str, (datetime.date, datetime.datetime)):
            return date_str
        try:
            return datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return None

    def get_transactions(self):
        """
        Carrega todas as transações da aba 'Financeiro' (linhas 7 a 509+)
        Usa o número da linha como ID único da transação.
        """
        wb = self.get_workbook(data_only=True)
        sheet = wb["Financeiro"]
        transactions = []
        
        today = datetime.date.today()

        for r in range(7, sheet.max_row + 1):
            # Interrompe se encontrar a linha de Balanço Geral
            # (que está após a tabela, por volta da linha 512+)
            desc_cell = sheet.cell(row=r, column=3).value
            dcto_cell = sheet.cell(row=r, column=2).value
            
            # Se for a linha do "Balanço Geral"
            if desc_cell == "Balanço Geral" or (dcto_cell is None and desc_cell == "Mês"):
                break
                
            # Verifica se há transação (se descrição ou doc está preenchido)
            # Para evitar carregar linhas vazias da tabela
            has_data = (desc_cell is not None and str(desc_cell).strip() != '') or \
                       (dcto_cell is not None and str(dcto_cell).strip() != '')
                       
            if has_data:
                # Lê colunas B a K (dados inseridos)
                n_dcto = self._clean_dcto(sheet.cell(row=r, column=2).value)
                descricao = sheet.cell(row=r, column=3).value
                data_vcto = self._parse_date(sheet.cell(row=r, column=4).value)
                data_pgto = self._parse_date(sheet.cell(row=r, column=5).value)
                parcela = sheet.cell(row=r, column=6).value
                entradas = sheet.cell(row=r, column=7).value or 0.0
                saidas = sheet.cell(row=r, column=8).value or 0.0
                conta = sheet.cell(row=r, column=9).value
                grupo = sheet.cell(row=r, column=10).value
                subgrupo = sheet.cell(row=r, column=11).value
                
                # Conversões de tipo numérico
                try: entradas = float(entradas)
                except ValueError: entradas = 0.0
                
                try: saidas = float(saidas)
                except ValueError: saidas = 0.0

                # Recalcula dinamicamente as colunas de status para consistência na web
                status = "Realizado" if data_pgto else "Previsto"
                
                vencida = "No prazo"
                if status == "Previsto" and data_vcto:
                    vcto_date = self._to_date_obj(data_vcto)
                    if vcto_date and vcto_date < today:
                        vencida = "Vencida"
                
                mes_pgto = None
                if data_pgto:
                    pgto_date = self._to_date_obj(data_pgto)
                    if pgto_date:
                        mes_pgto = pgto_date.month

                transactions.append({
                    "id": r, # Número da linha como ID único
                    "n_dcto": n_dcto,
                    "descricao": descricao,
                    "data_vcto": data_vcto,
                    "data_pgto": data_pgto,
                    "parcela": parcela,
                    "entradas": entradas,
                    "saidas": saidas,
                    "conta": conta,
                    "grupo": grupo,
                    "subgrupo": subgrupo,
                    "status": status,
                    "vencida": vencida,
                    "mes": mes_pgto
                })
                
        wb.close()
        return transactions

    def write_formulas(self, sheet, r):
        """Escreve as fórmulas padrão do Excel nas colunas L a V para a linha 'r'"""
        # STATUS (L)
        sheet.cell(row=r, column=12, value=f'=IF(E{r}="","Previsto","Realizado")')
        # VENCIDA (M)
        sheet.cell(row=r, column=13, value=f'=IF(AND(L{r}="Previsto",D{r}<TODAY()),"Vencida","No prazo")')
        # MÊS (N)
        sheet.cell(row=r, column=14, value=f'=IF(E{r}="","",MONTH(E{r}))')
        
        # Colunas O a R (Entradas janelas: 7d, 14d, 30d, todos)
        sheet.cell(row=r, column=15, value=f'=IF(AND(D{r}>=TODAY(),D{r}<=TODAY()+7),G{r},0)')
        sheet.cell(row=r, column=16, value=f'=IF(AND(D{r}>=TODAY(),D{r}<=TODAY()+14),G{r},0)')
        sheet.cell(row=r, column=17, value=f'=IF(AND(D{r}>=TODAY(),D{r}<=TODAY()+30),G{r},0)')
        sheet.cell(row=r, column=18, value=f'=IF(AND(D{r}>=TODAY(),D{r}<=TODAY()+10000),G{r},0)')
        
        # Colunas S a V (Saídas janelas: 7d, 14d, 30d, todos)
        sheet.cell(row=r, column=19, value=f'=IF(AND(D{r}>=TODAY(),D{r}<=TODAY()+7),H{r},0)')
        sheet.cell(row=r, column=20, value=f'=IF(AND(D{r}>=TODAY(),D{r}<=TODAY()+14),H{r},0)')
        sheet.cell(row=r, column=21, value=f'=IF(AND(D{r}>=TODAY(),D{r}<=TODAY()+30),H{r},0)')
        sheet.cell(row=r, column=22, value=f'=IF(AND(D{r}>=TODAY(),D{r}<=TODAY()+10000),H{r},0)')

    def add_transactions(self, new_txs):
        """
        Adiciona múltiplas transações no Excel.
        Busca a primeira linha vazia entre 7 e 509.
        Se exceder, estende a tabela.
        """
        wb = self.get_workbook(data_only=False)
        sheet = wb["Financeiro"]
        
        added_count = 0
        current_row = 7
        
        for tx in new_txs:
            # Encontra a próxima linha vazia onde DESCRIÇÃO (col C) e Nº DCTO (col B) são None
            while True:
                # Evita sobrescrever o Balanço Geral (linha 512+)
                desc_val = sheet.cell(row=current_row, column=3).value
                dcto_val = sheet.cell(row=current_row, column=2).value
                
                # Se encontrarmos a linha de Balanço Geral, precisamos empurrá-la ou inserir linhas
                if desc_val == "Balanço Geral" or (dcto_val is None and desc_val == "Mês"):
                    # Insere uma linha imediatamente antes do Balanço Geral
                    sheet.insert_rows(current_row, 1)
                    # Escreve as fórmulas para essa nova linha criada
                    self.write_formulas(sheet, current_row)
                    break
                
                # Se a linha atual estiver vazia
                if (desc_val is None or str(desc_val).strip() == '') and \
                   (dcto_val is None or str(dcto_val).strip() == ''):
                    break
                
                current_row += 1
            
            # Escreve dados na linha encontrada/criada
            # B: Nº DCTO, C: DESCRIÇÃO, D: DATA VCTO, E: DATA PGTO, F: PARCELA, G: ENTRADAS, H: SAÍDAS, I: CONTA, J: GRUPO, K: SUBGRUPO
            c_dcto = sheet.cell(row=current_row, column=2)
            c_dcto.value = tx.get("n_dcto")
            c_dcto.number_format = 'General'
            
            c_desc = sheet.cell(row=current_row, column=3)
            c_desc.value = tx.get("descricao")
            c_desc.number_format = 'General'
            
            sheet.cell(row=current_row, column=4).value = self._to_date_obj(tx.get("data_vcto"))
            sheet.cell(row=current_row, column=5).value = self._to_date_obj(tx.get("data_pgto"))
            
            c_part = sheet.cell(row=current_row, column=6)
            c_part.value = tx.get("parcela")
            c_part.number_format = 'General'
            
            # Entradas e saídas numéricas
            entradas = tx.get("entradas")
            saidas = tx.get("saidas")
            sheet.cell(row=current_row, column=7).value = float(entradas) if entradas else None
            sheet.cell(row=current_row, column=8).value = float(saidas) if saidas else None
            
            c_conta = sheet.cell(row=current_row, column=9)
            c_conta.value = tx.get("conta")
            c_conta.number_format = 'General'
            
            c_grupo = sheet.cell(row=current_row, column=10)
            c_grupo.value = tx.get("grupo")
            c_grupo.number_format = 'General'
            
            c_subg = sheet.cell(row=current_row, column=11)
            c_subg.value = tx.get("subgrupo")
            c_subg.number_format = 'General'
            
            # Garante que a linha possui as fórmulas, caso tenha sido criada fora do padrão original
            # ou para limpar qualquer resquício
            self.write_formulas(sheet, current_row)
            
            added_count += 1
            current_row += 1
            
        wb.save(self.filepath)
        wb.close()
        return added_count

    def update_transaction(self, tx_id, tx_data):
        """
        Atualiza uma transação existente com base no ID (número da linha).
        """
        wb = self.get_workbook(data_only=False)
        sheet = wb["Financeiro"]
        
        r = int(tx_id)
        if r < 7 or r > sheet.max_row:
            wb.close()
            raise ValueError(f"ID/Linha inválida para atualização: {tx_id}")
            
        c_dcto = sheet.cell(row=r, column=2)
        c_dcto.value = tx_data.get("n_dcto")
        c_dcto.number_format = 'General'
        
        c_desc = sheet.cell(row=r, column=3)
        c_desc.value = tx_data.get("descricao")
        c_desc.number_format = 'General'
        
        sheet.cell(row=r, column=4).value = self._to_date_obj(tx_data.get("data_vcto"))
        sheet.cell(row=r, column=5).value = self._to_date_obj(tx_data.get("data_pgto"))
        
        c_part = sheet.cell(row=r, column=6)
        c_part.value = tx_data.get("parcela")
        c_part.number_format = 'General'
        
        entradas = tx_data.get("entradas")
        saidas = tx_data.get("saidas")
        sheet.cell(row=r, column=7).value = float(entradas) if entradas else None
        sheet.cell(row=r, column=8).value = float(saidas) if saidas else None
        
        c_conta = sheet.cell(row=r, column=9)
        c_conta.value = tx_data.get("conta")
        c_conta.number_format = 'General'
        
        c_grupo = sheet.cell(row=r, column=10)
        c_grupo.value = tx_data.get("grupo")
        c_grupo.number_format = 'General'
        
        c_subg = sheet.cell(row=r, column=11)
        c_subg.value = tx_data.get("subgrupo")
        c_subg.number_format = 'General'
        
        # Reescreve as fórmulas para garantir
        self.write_formulas(sheet, r)
        
        wb.save(self.filepath)
        wb.close()
        return True

    def delete_transaction(self, tx_id):
        """
        Exclui uma transação limpando as colunas B a K na linha especificada.
        Garante que a linha permaneça vazia para futura inserção sem corromper as fórmulas.
        """
        wb = self.get_workbook(data_only=False)
        sheet = wb["Financeiro"]
        
        r = int(tx_id)
        if r < 7 or r > sheet.max_row:
            wb.close()
            raise ValueError(f"ID/Linha inválida para exclusão: {tx_id}")
            
        # Limpa as colunas de dados B a K
        for col in range(2, 12):
            sheet.cell(row=r, column=col).value = None
            
        # Reseta as fórmulas para que reflitam a linha limpa (STATUS Previsto, etc)
        self.write_formulas(sheet, r)
        
        wb.save(self.filepath)
        wb.close()
        return True

    def get_reports(self):
        """
        Calcula os relatórios por conta:
        - Crédito total (CT)
        - Débito total (DT)
        - Saldo total (ST) = CT - DT
        """
        transactions = self.get_transactions()
        reports = {}
        
        for tx in transactions:
            conta = tx["conta"]
            if not conta or str(conta).strip() == '':
                conta = "Não informada"
            conta = str(conta).strip()
            
            if conta not in reports:
                reports[conta] = {"credit": 0.0, "debit": 0.0, "balance": 0.0}
                
            reports[conta]["credit"] += tx["entradas"]
            reports[conta]["debit"] += tx["saidas"]
            
        for conta in reports:
            reports[conta]["balance"] = reports[conta]["credit"] - reports[conta]["debit"]
            
        return reports

    def get_dashboard_data(self):
        """
        Retorna dados consolidados para o Dashboard:
        - KPIs superiores (Contas vencendo hoje, atrasadas, recebimentos hoje, recebimentos atrasados)
        - Gráficos de gasto mensal e anual por subgrupo (TTGM e TTGA)
        - Projeção de fluxo de caixa futuro
        """
        transactions = self.get_transactions()
        today = datetime.date.today()
        
        kpis = {
            "contas_vencendo_hoje": 0.0,
            "contas_atrasadas": 0.0,
            "recebimentos_vencendo_hoje": 0.0,
            "recebimentos_atrasados": 0.0
        }
        
        ttgm = {} # { "ANO_MES": { "subgrupo": total } }
        ttga = {} # { "ANO": { "subgrupo": total } }
        
        # Fluxo de caixa previsto por data prevista
        cash_flow_timeline = {} # { "data": { "entradas": X, "saidas": Y } }
        
        for tx in transactions:
            status = tx["status"]
            vencida = tx["vencida"]
            
            data_vcto_obj = self._to_date_obj(tx["data_vcto"])
            data_pgto_obj = self._to_date_obj(tx["data_pgto"])
            
            # Data Prevista: data_pgto se pago, senão data_vcto
            data_prevista_obj = data_pgto_obj if data_pgto_obj else data_vcto_obj
            data_prevista_str = data_prevista_obj.strftime("%Y-%m-%d") if data_prevista_obj else None
            
            # --- CÁLCULO DE KPIS ---
            if status == "Previsto" and data_vcto_obj:
                # Contas (Saídas)
                if tx["saidas"] > 0:
                    if data_vcto_obj == today:
                        kpis["contas_vencendo_hoje"] += tx["saidas"]
                    if vencida == "Vencida":
                        kpis["contas_atrasadas"] += tx["saidas"]
                # Recebimentos (Entradas)
                if tx["entradas"] > 0:
                    if data_vcto_obj == today:
                        kpis["recebimentos_vencendo_hoje"] += tx["entradas"]
                    if vencida == "Vencida":
                        kpis["recebimentos_atrasados"] += tx["entradas"]
                        
            # --- CÁLCULO DE GASTOS (SAÍDAS) POR SUBGRUPO ---
            if tx["saidas"] > 0 and tx["subgrupo"] and data_prevista_obj:
                subg = tx["subgrupo"].strip()
                year = data_prevista_obj.year
                month = data_prevista_obj.month
                
                # Anual (TTGA)
                y_str = str(year)
                if y_str not in ttga:
                    ttga[y_str] = {}
                ttga[y_str][subg] = ttga[y_str].get(subg, 0.0) + tx["saidas"]
                
                # Mensal (TTGM)
                ym_str = f"{year}-{month:02d}"
                if ym_str not in ttgm:
                    ttgm[ym_str] = {}
                ttgm[ym_str][subg] = ttgm[ym_str].get(subg, 0.0) + tx["saidas"]
                
            # --- CÁLCULO DE CRONOGRAMA DE CAIXA PREVISTO ---
            if data_prevista_str:
                if data_prevista_str not in cash_flow_timeline:
                    cash_flow_timeline[data_prevista_str] = {"entradas": 0.0, "saidas": 0.0}
                cash_flow_timeline[data_prevista_str]["entradas"] += tx["entradas"]
                cash_flow_timeline[data_prevista_str]["saidas"] += tx["saidas"]

        # Formata timeline ordenando por data
        timeline_sorted = []
        for dt_str in sorted(cash_flow_timeline.keys()):
            timeline_sorted.append({
                "date": dt_str,
                "entradas": cash_flow_timeline[dt_str]["entradas"],
                "saidas": cash_flow_timeline[dt_str]["saidas"]
            })
            
        return {
            "kpis": kpis,
            "ttgm": ttgm,
            "ttga": ttga,
            "timeline": timeline_sorted
        }
