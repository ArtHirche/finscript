import os
import shutil
import openpyxl
from backend.excel_manager import ExcelManager, DEFAULT_EXCEL_PATH

TEST_EXCEL_PATH = r"c:\projetos\finscript\Dashboard Ocimar 2023_diagnostic_fmt.xlsx"

if os.path.exists(TEST_EXCEL_PATH):
    os.remove(TEST_EXCEL_PATH)

shutil.copy(DEFAULT_EXCEL_PATH, TEST_EXCEL_PATH)

# Write with format = General
wb = openpyxl.load_workbook(TEST_EXCEL_PATH)
sheet = wb["Financeiro"]
cell = sheet.cell(row=7, column=2)
cell.value = 101
cell.number_format = 'General'
wb.save(TEST_EXCEL_PATH)
wb.close()

# Read back
wb_read = openpyxl.load_workbook(TEST_EXCEL_PATH, data_only=True)
sheet_read = wb_read["Financeiro"]
val = sheet_read.cell(row=7, column=2).value
print(f"Read back value: {val}, type: {type(val)}")
wb_read.close()

if os.path.exists(TEST_EXCEL_PATH):
    os.remove(TEST_EXCEL_PATH)
