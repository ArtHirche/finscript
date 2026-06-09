import os
import shutil
import openpyxl
from backend.excel_manager import ExcelManager, DEFAULT_EXCEL_PATH

TEST_EXCEL_PATH = r"c:\projetos\finscript\Dashboard Ocimar 2023_diagnostic_del.xlsx"

if os.path.exists(TEST_EXCEL_PATH):
    os.remove(TEST_EXCEL_PATH)

shutil.copy(DEFAULT_EXCEL_PATH, TEST_EXCEL_PATH)
manager = ExcelManager(TEST_EXCEL_PATH)

# Add
test_txs = [{
    "n_dcto": 301,
    "descricao": "Deletar",
    "data_vcto": "2026-06-20",
    "data_pgto": None,
    "parcela": "1/1",
    "entradas": 100.0,
    "saidas": 0.0,
    "conta": "Banco",
    "grupo": "RECEITAS_OPERACIONAIS",
    "subgrupo": "Receitas de Vendas"
}]
manager.add_transactions(test_txs)

txs = manager.get_transactions()
tx = [t for t in txs if t["n_dcto"] == 301][0]
tx_id = tx["id"]
print(f"Added at line {tx_id}. Confirming from get_transactions: n_dcto={tx['n_dcto']}, desc={tx['descricao']}")

# Let's inspect raw cells of line tx_id using openpyxl directly before delete
wb = openpyxl.load_workbook(TEST_EXCEL_PATH)
sheet = wb["Financeiro"]
print("Raw cell values in file before delete:")
for col in range(2, 12):
    print(f"Col {col}: {sheet.cell(row=tx_id, column=col).value}")
wb.close()

# Delete
manager.delete_transaction(tx_id)

# Inspect raw cells after delete
wb = openpyxl.load_workbook(TEST_EXCEL_PATH)
sheet = wb["Financeiro"]
print("Raw cell values in file after delete:")
for col in range(2, 12):
    print(f"Col {col}: {sheet.cell(row=tx_id, column=col).value}")
wb.close()

# Clean up
if os.path.exists(TEST_EXCEL_PATH):
    os.remove(TEST_EXCEL_PATH)
